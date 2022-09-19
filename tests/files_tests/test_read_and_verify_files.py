import time
import csv
from helpers.files_funcs import generate_absolute_path, add_files_to_zip, extract_all_files_from_zip
from openpyxl import load_workbook
from PyPDF2 import PdfReader


def test_verify_archive_files():
    """Packing archive"""

    time_stamp = str(time.time()).replace('.', '_')
    archive_path = f'{generate_absolute_path("resources")}/archive_{time_stamp}.zip'
    add_files_to_zip(generate_absolute_path('resources'), archive_path, ['csv', 'xlsx', 'pdf'])

    """Extracting files"""

    extract_path = f'{generate_absolute_path("resources")}/archive_{time_stamp}'
    extract_all_files_from_zip(extract_path, archive_path)

    """Checking csv file"""

    cities_without_kladr = []

    with open(f'{extract_path}/cities.csv') as csvfile:
        cities = csv.DictReader(csvfile)
        for row in cities:
            if len(row['kladr']) == 0:
                cities_without_kladr.append(row["name"])

    if len(cities_without_kladr) > 0:
        with open(f'{extract_path}/cities_without_kladr.txt', 'a') as f:
            for city in cities_without_kladr:
                f.write(f'{city}\n')

    """Checking xlsx file"""

    wb = load_workbook(f'{extract_path}/employees.xlsx')
    ws = wb.active

    female_number = 0
    male_number = 0

    for cell in ws['D']:
        if cell.value == 'Female':
            female_number += 1
        elif cell.value == 'Female':
            male_number += 1

    assert female_number >= male_number

    """Checking pdf file"""

    pdf_reader = PdfReader(f'{extract_path}/file-example_PDF_500_kB.pdf')
    page = pdf_reader.pages[0]
    text = page.extract_text()

    assert text.startswith('Lorem ipsum') is True
